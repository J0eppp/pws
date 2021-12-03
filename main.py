from argparse import ArgumentParser
import time

import solver.utils as utils
from solver.parser import parse_json_file
from solver.datatypes import Timetable

SEPERATION_STRING = "-==================================-"


def main():
    parser = ArgumentParser(description="Process arguments for this program")
    parser.add_argument("-s", "--solver", type=str,
                        help="Select the linear programming solver")
    parser.add_argument("-d", "--data", type=str, help="Select the JSON file")
    parser.add_argument("--save", type=str, help="Where to save the model")
    parser.add_argument("--display", action="store_true",
                        help="Display the graph (when using the GC sovler)")
    parser.add_argument("-g", "--gui", action="store_true", help="Use the GUI")
    parser.add_argument("-v", "--verbosity", type=int,
                        help="Set the verbosity")
    parser.add_argument("--excel", type=str,
                        help="Save the result as an Excel document")
    args = parser.parse_args()
    data_file = args.data
    save_file = args.save

    # Parse the CLI arguments
    if args.gui == False:
        if data_file == None:
            utils.uprint(
                "Please specify an input file with the -d/--data argument")
            return

        if args.solver == None:
            utils.uprint("Please specify a solving method")
            utils.uprint("Available solving methods:")
            utils.uprint("    1. Linear programming (lp)")
            utils.uprint("    2. Graph colouring (gc)")
            utils.uprint("    3. Random solver (rs)")
            return

        start_time = 0
        end_time = 0

        # Read and parse the file with all the data
        if args.verbosity == 1:
            utils.uprint(SEPERATION_STRING)
            utils.uprint("Reading and parsing the file....")
            start_time = time.time()

        timetable: Timetable = parse_json_file(data_file)

        if args.verbosity == 1:
            end_time = time.time()
            utils.uprint("Done reading and parsing the file")
            utils.uprint(
                f"Reading and parsing the file took {end_time - start_time} seconds")
            utils.uprint("Loaded {amount_of_groups} group{multiple_groups} and {amount_of_teachers} teacher{multiple_teachers}".format(amount_of_groups=len(
                timetable.groups), amount_of_teachers=len(timetable.teachers), multiple_groups="s" if len(timetable.groups) > 1 else "", multiple_teachers="s" if len(timetable.teachers) > 1 else ""))
            utils.uprint(SEPERATION_STRING)

        solver = None

        # Select solver
        if args.solver == "lp":
            from solver.LPSolver import LPSolver
            verbosity = args.verbosity if args.verbosity != None else 0
            solver = LPSolver(timetable, verbosity, args.save)
        elif args.solver == "gc":
            from solver.GCSolver import GCSolver
            display = False
            if args.display != None:
                display = args.display
            solver = GCSolver(timetable, display=display, save=save_file)
        elif args.solver == "rs":
            from solver.RandomSolver import RandomSolver
            verbosity = args.verbosity if args.verbosity != None else 0
            solver = RandomSolver(timetable, verbosity=verbosity)

        timetable = solver.solve()

        if args.excel != None:
            # Save the document as an Excel document
            from openpyxl import Workbook
            if args.verbosity == 1:
                utils.uprint(SEPERATION_STRING)
                utils.uprint("Saving the timetable as an Excel document")
                start_time = time.time()
            workbook = Workbook()
            worksheet = workbook.active
            for i in range(len(timetable.groups)):
                group = timetable.groups[i]
                column = i + 1
                worksheet.cell(column=column, row=1).value = group.name
                days = [[] for _ in range(timetable.amount_of_days_a_week)]
                for lesson in group.lessons:
                    days[lesson.day].append(lesson)

                [day.sort(key=lambda x: x.hour) for day in days]
                row = 2
                for day in days:
                    for lesson in day:
                        worksheet.cell(
                            row=row, column=column).value = lesson.excel_str()
                        row += 1

            workbook.save(args.excel)
            if args.verbosity == 1:
                end_time = time.time()
                utils.uprint(
                    "Saving the timetable as an Excel document is done")
                utils.uprint(
                    f"Saving the document took {end_time - start_time} seconds")
                utils.uprint(SEPERATION_STRING)

    if args.gui == True:
        from solver.GUI import GUI
        gui = GUI(args)
        gui.run()


if __name__ == "__main__":
    start_time = time.time()
    main()
    end_time = time.time()
    utils.uprint(SEPERATION_STRING)
    utils.uprint(f"Total execution time is {end_time - start_time} seconds")
    utils.uprint(SEPERATION_STRING)
